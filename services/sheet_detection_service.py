from openpyxl import load_workbook

from utils.text_utils import normalize_key


def detect_workbook_sheets(file_path: str, mappings: dict):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    available_sheet_names = workbook.sheetnames

    question_sheet_name = match_sheet_name(
        available_sheet_names,
        mappings["sheet_detection"]["question_sheet_candidates"],
    )
    data_sheet_name = match_sheet_name(
        available_sheet_names,
        mappings["sheet_detection"]["data_sheet_candidates"],
    )
    if not question_sheet_name or not data_sheet_name:
        raise ValueError(f"Required sheets not found. Available sheets: {available_sheet_names}")

    return {
        "sheet_names": available_sheet_names,
        "question_sheet_name": question_sheet_name,
        "data_sheet_name": data_sheet_name,
    }


def match_sheet_name(available_sheet_names, candidates):
    normalized_candidates = {normalize_key(candidate): candidate for candidate in candidates}
    for sheet_name in available_sheet_names:
        if normalize_key(sheet_name) in normalized_candidates:
            return sheet_name
    return None
