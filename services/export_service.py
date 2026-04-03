from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side, Alignment

from services.file_service import build_output_paths


def export_parse_outputs(upload_id: int, parse_result: dict, app_config: dict):
    output_paths = build_output_paths(upload_id, app_config["OUTPUT_FOLDER"])

    parse_result["respondent_data_clean"].to_excel(output_paths["cleaned_data_path"], index=False)
    parse_result["respondent_data_clean"].to_pickle(output_paths["cleaned_data_pickle_path"])
    pd.DataFrame(parse_result["question_dictionary"]).to_excel(output_paths["question_dictionary_path"], index=False)

    with pd.ExcelWriter(output_paths["mapping_audit_path"], engine="openpyxl") as writer:
        pd.DataFrame(parse_result["variable_catalog"]).to_excel(writer, sheet_name="variable_catalog", index=False)
        pd.DataFrame(parse_result["question_dictionary"]).to_excel(writer, sheet_name="question_dictionary", index=False)

    return {key: str(Path(path)) for key, path in output_paths.items() if key != "run_dir"}


def export_quota_dashboard(upload_id: int, dashboard_payload: dict, app_config: dict):
    output_paths = build_output_paths(upload_id, app_config["OUTPUT_FOLDER"])
    file_path = output_paths["quota_dashboard_path"]

    if dashboard_payload.get("report_type") == "banner_table":
        return export_banner_dashboard(file_path, dashboard_payload)

    metadata_rows = [
        ["Source file name", dashboard_payload["source_file_name"]],
        ["Processing timestamp", dashboard_payload["processed_at"].strftime("%Y-%m-%d %H:%M:%S") if dashboard_payload.get("processed_at") else ""],
        ["Horizontal variable code", dashboard_payload["horizontal_code"]],
        ["Horizontal variable label", dashboard_payload["horizontal_label"]],
        ["Vertical variable code", dashboard_payload["vertical_code"]],
        ["Vertical variable label", dashboard_payload["vertical_label"]],
        ["Display mode", dashboard_payload["display_mode"]],
        ["Percent mode", dashboard_payload["percent_mode"]],
        ["Accepted valid base size", dashboard_payload["accepted_base_size"]],
        [],
    ]

    table_views = dashboard_payload.get("table_views") or [
        {
            "mode": dashboard_payload["display_mode"],
            "rows": dashboard_payload["rows"],
            "total_row": dashboard_payload["total_row"],
        }
    ]
    table_header = [dashboard_payload["vertical_label"], *dashboard_payload["column_categories"], "Total"]
    table_rows = []
    table_start_rows = []
    current_row = len(metadata_rows) + 1

    for index, table_view in enumerate(table_views):
        table_start_rows.append(current_row + 1)
        table_rows.append([f"{table_view['mode']} table"])
        table_rows.append(table_header)
        for row in table_view["rows"]:
            table_rows.append([row["label"], *[cell["display"] for cell in row["cells"]], row["total"]["display"]])
        table_rows.append(
            [
                table_view["total_row"]["label"],
                *[cell["display"] for cell in table_view["total_row"]["cells"]],
                table_view["total_row"]["total"]["display"],
            ]
        )
        current_row += len(table_view["rows"]) + 3
        if index < len(table_views) - 1:
            table_rows.append([])
            current_row += 1

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        pd.DataFrame(metadata_rows + table_rows).to_excel(writer, sheet_name="Quota Dashboard", header=False, index=False)

    format_quota_dashboard_export(file_path, table_start_rows=table_start_rows, table_width=len(table_header))

    return str(file_path)


def format_quota_dashboard_export(file_path: str, table_start_rows: list[int], table_width: int):
    workbook = load_workbook(file_path)
    worksheet = workbook["Quota Dashboard"]

    bold_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    total_fill = PatternFill("solid", fgColor="EDEDED")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    first_table_start_row = table_start_rows[0] if table_start_rows else 1
    for row in range(1, first_table_start_row):
        worksheet.cell(row=row, column=1).font = bold_font

    last_col = table_width
    for table_start_row in table_start_rows:
        title_cell = worksheet.cell(row=table_start_row, column=1)
        title_cell.font = bold_font

        header_row = table_start_row + 1
        for column in range(1, table_width + 1):
            cell = worksheet.cell(row=header_row, column=column)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        row = header_row + 1
        while row <= worksheet.max_row and normalize_excel_value(worksheet.cell(row=row, column=1).value) not in {"", "count table", "percent table"}:
            for column in range(1, last_col + 1):
                cell = worksheet.cell(row=row, column=column)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            worksheet.cell(row=row, column=last_col).fill = total_fill
            worksheet.cell(row=row, column=last_col).font = bold_font
            row += 1
        total_row = row - 1
        for column in range(1, last_col + 1):
            worksheet.cell(row=total_row, column=column).fill = total_fill
            worksheet.cell(row=total_row, column=column).font = bold_font

    worksheet.freeze_panes = worksheet.cell(row=first_table_start_row + 2, column=2)

    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    workbook.save(file_path)


def normalize_excel_value(value):
    return str(value or "").strip().lower()


def export_banner_dashboard(file_path: str, dashboard_payload: dict):
    banner_tree_text = " > ".join(
        [item.get("variable_code", "") for item in dashboard_payload.get("selected_banner_tree", []) if item.get("variable_code")]
    )
    metadata_rows = [
        ["Source file name", dashboard_payload["source_file_name"]],
        ["Processing timestamp", dashboard_payload["processed_at"].strftime("%Y-%m-%d %H:%M:%S") if dashboard_payload.get("processed_at") else ""],
        ["Row questions", ", ".join(dashboard_payload.get("selected_row_variables", []))],
        ["Banner columns", ", ".join(dashboard_payload.get("selected_banner_variables", []))],
        ["Banner layout", dashboard_payload.get("selected_banner_layout_mode", "flat")],
        ["Header tree", banner_tree_text],
        ["Display mode", dashboard_payload["display_mode"]],
        ["Percent mode", dashboard_payload["percent_mode"]],
        ["Accepted valid base size", dashboard_payload["accepted_base_size"]],
        [],
    ]

    rows = []
    start_rows = []
    current_row = len(metadata_rows) + 1
    for index, banner_view in enumerate(dashboard_payload.get("banner_views", [])):
        start_rows.append(current_row + 1)
        rows.append([f"{banner_view['mode']} table"])
        rows.append(["Question", "Category", *[column["label"] for group in dashboard_payload["column_groups"] for column in group["columns"]]])
        rows.append(["", "", *[group["question_label"] for group in dashboard_payload["column_groups"] for _ in group["columns"]]])
        for section in banner_view["sections"]:
            first = True
            for row in section["rows"]:
                rows.append([
                    f"{section['row_variable']} - {section['question_label']}" if first else "",
                    row["category_label"],
                    *[cell["display"] for cell in row["cells"]],
                ])
                first = False
        current_row += 3 + sum(len(section["rows"]) for section in banner_view["sections"])
        if index < len(dashboard_payload.get("banner_views", [])) - 1:
            rows.append([])
            current_row += 1

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        pd.DataFrame(metadata_rows + rows).to_excel(writer, sheet_name="Quota Dashboard", header=False, index=False)

    workbook = load_workbook(file_path)
    worksheet = workbook["Quota Dashboard"]
    bold_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in range(1, len(metadata_rows) + 1):
        worksheet.cell(row=row, column=1).font = bold_font
    for start_row in start_rows:
        worksheet.cell(row=start_row, column=1).font = bold_font
        for header_row in (start_row + 1, start_row + 2):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if cell.row >= len(metadata_rows) + 1:
                cell.border = thin_border
    worksheet.freeze_panes = "C11"
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)
    workbook.save(file_path)
    return str(file_path)
