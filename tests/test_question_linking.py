from openpyxl import Workbook

from services.question_parser_service import parse_question_sheet
from services.raw_data_parser_service import parse_data_sheet


def build_workbook(path):
    workbook = Workbook()
    question_sheet = workbook.active
    question_sheet.title = "Question"
    question_sheet.append(["Name of items", "Question type", "Question(Matrix)", "Question(Normal)", "1", "2", "3"])
    question_sheet.append(["S2", "SA", "", "Distance", "0 - 2.49 km", "2.5 - 5 km", "5.1 - 7 km"])
    question_sheet.append(["S3", "SA", "", "Direction", "Bac", "Dong", "Tay"])

    data_sheet = workbook.create_sheet("Data")
    for _ in range(4):
        data_sheet.append(["", "", ""])
    data_sheet.append(["Tu choi", "S2", "S3"])
    data_sheet.append(["", "", ""])
    data_sheet.append(["", "", ""])
    data_sheet.append(["", 1, 1])
    data_sheet.append(["x", 2, 2])
    workbook.save(path)


def test_question_linking_and_decoding(tmp_path):
    file_path = tmp_path / "linked.xlsx"
    build_workbook(file_path)
    mappings = {
        "parser_options": {"question_header_row": 1, "data_header_row": 5, "data_start_row": 8},
        "rejection_column_candidates": ["Tu choi"],
    }
    question_result = parse_question_sheet(str(file_path), "Question", mappings)
    data_result = parse_data_sheet(
        str(file_path),
        "Data",
        upload_run_id=1,
        source_file_name="linked.xlsx",
        mappings=mappings,
        question_metadata=question_result["question_metadata"],
    )
    df = data_result["respondent_data_clean"]
    assert df.loc[0, "decoded__S2"] == "0 - 2.49 km"
    assert df.loc[0, "decoded__S3"] == "Bac"
    assert df.loc[1, "is_rejected"] == True  # noqa: E712


def test_duplicate_data_headers_are_made_unique(tmp_path):
    file_path = tmp_path / "linked_dup.xlsx"
    workbook = Workbook()
    question_sheet = workbook.active
    question_sheet.title = "Question"
    question_sheet.append(["Name of items", "Question type", "Question(Matrix)", "Question(Normal)", "1", "2"])
    question_sheet.append(["S2", "SA", "", "Distance", "0 - 2.49 km", "2.5 - 5 km"])

    data_sheet = workbook.create_sheet("Data")
    for _ in range(4):
        data_sheet.append(["", "", ""])
    data_sheet.append(["Tu choi", "S2", "S2"])
    data_sheet.append(["", "", ""])
    data_sheet.append(["", "", ""])
    data_sheet.append(["", 1, 2])
    workbook.save(file_path)

    mappings = {
        "parser_options": {"question_header_row": 1, "data_header_row": 5, "data_start_row": 8},
        "rejection_column_candidates": ["Tu choi"],
    }
    question_result = parse_question_sheet(str(file_path), "Question", mappings)
    data_result = parse_data_sheet(
        str(file_path),
        "Data",
        upload_run_id=1,
        source_file_name="linked_dup.xlsx",
        mappings=mappings,
        question_metadata=question_result["question_metadata"],
    )
    df = data_result["respondent_data_clean"]
    assert "decoded__S2" in df.columns
    assert "decoded__S2_dup2" in df.columns
    assert df.loc[0, "decoded__S2"] == "0 - 2.49 km"
    assert df.loc[0, "decoded__S2_dup2"] == "2.5 - 5 km"
